from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

F = lambda **k: Font(name='Arial', size=10, **k)
BLUE, BLACK, GREEN = Font(name='Arial', size=10, color='0000FF'), Font(name='Arial', size=10), Font(name='Arial', size=10, color='008000')
HDR = Font(name='Arial', size=11, bold=True, color='FFFFFF')
SEC = Font(name='Arial', size=10, bold=True)
TITLE = Font(name='Arial', size=13, bold=True)
FILL_HDR = PatternFill('solid', start_color='2F5233')
FILL_SEC = PatternFill('solid', start_color='E8F0E8')
FILL_KEY = PatternFill('solid', start_color='FFFF00')
KRW = '#,##0;(#,##0);"-"'
PCT = '0.0%;(0.0%);"-"'
NUM = '#,##0.0;(#,##0.0);"-"'
INT = '#,##0;(#,##0);"-"'
MULT = '0.0"x"'

NM = 36  # months
C0 = 3   # first month column (C)
def mc(i): return get_column_letter(C0 + i - 1)  # month i (1-based) -> col letter

wb = Workbook()
OUT_DIR = Path(__file__).resolve().parent / 'out'
OUT_DIR.mkdir(parents=True, exist_ok=True)

# ============ Sheet 1: 가정 ============
A = wb.active; A.title = '가정'
A.column_dimensions['A'].width = 34
for c in 'BCDE': A.column_dimensions[c].width = 13
A.column_dimensions['F'].width = 60

def put(ws, cell, val, font=BLACK, fmt=None, fill=None, align=None):
    ws[cell] = val; ws[cell].font = font
    if fmt: ws[cell].number_format = fmt
    if fill: ws[cell].fill = fill
    if align: ws[cell].alignment = Alignment(horizontal=align)

put(A, 'A1', 'Kindy 재무모델 v1.0 — 가정 (Assumptions)', TITLE)
put(A, 'A2', '색상: 파랑=입력값(수정 가능) / 검정=수식 / 초록=타 시트 링크 / 노랑배경=핵심 가정', Font(name='Arial', size=9, italic=True, color='777777'))
put(A, 'A4', '시나리오 선택 (1=보수 / 2=기본 / 3=낙관)', SEC)
put(A, 'B4', 2, BLUE, INT, FILL_KEY)
put(A, 'C4', '보수', SEC, align='center'); put(A, 'D4', '기본', SEC, align='center'); put(A, 'E4', '낙관', SEC, align='center')

rows = [
 # (row, label, applied_formula_or_value, cons, base, opt, fmt, note)
 (6,  '── B2C 가격 ──', None, None, None, None, None, ''),
 (7,  '베타 월 구독료 (VAT포함, ₩)', 19000, None, None, None, KRW, '얼리버드: 첫 6개월 신규 대상. 근거: 백서 v1.1 §8.4'),
 (8,  '정가 월 구독료 (VAT포함, ₩)', 25000, None, None, None, KRW, '근거: IR ARPU 25,000 / 백서 스위트스팟'),
 (9,  '베타 가격 적용 종료월 (M)', 6, None, None, None, INT, 'M7부터 정가'),
 (10, '얼리버드 락인 ARPU 희석률', 0.08, None, None, None, PCT, '락인 가구 잔존으로 정가 대비 blended ARPU 하락분(추정)'),
 (11, '형제 프로필 attach율', 0.10, None, None, None, PCT, '가구당 자녀 추가 프로필'),
 (12, '형제 프로필 월가 (VAT포함, ₩)', 7000, None, None, None, KRW, ''),
 (13, '연간플랜 비중 (신규 중)', 0.0, None, None, None, PCT, '기본 0%. 활성화 시 현금 선수 효과가 현금 시트에 반영'),
 (14, '연간플랜 가격 (VAT포함, ₩)', 190000, None, None, None, KRW, '월 15,833 효과 = 17% 할인'),
 (16, '── 퍼널 (도서관 키오스크) ──', None, None, None, None, None, ''),
 (17, '키오스크 세션 수 / 관 / 월', 1200, None, None, None, INT, '일 40세션 × 30일. 실측치로 교체 필요(스프린트 W2 계측)'),
 (18, '세션→QR 스캔율', '=CHOOSE($B$4,C18,D18,E18)', 0.04, 0.06, 0.08, PCT, '★ 최대 불확실 변수. W2 실측으로 교체'),
 (19, 'QR→가입(부모동의) 완료율', 0.45, None, None, None, PCT, 'PIPA 동의 플로우 포함'),
 (20, '가입→무료체험(3세션) 시작률', 0.60, None, None, None, PCT, ''),
 (21, '체험→유료 전환율', '=CHOOSE($B$4,C21,D21,E21)', 0.35, 0.45, 0.55, PCT, '근거: 온보딩 리서치 트라이얼→유료 중앙값 50~60%'),
 (22, '월 이탈률 (churn)', '=CHOOSE($B$4,C22,D22,E22)', 0.35, 0.25, 0.15, PCT, '근거: 백서 v1.1 §8.1'),
 (24, '── Paid 마케팅 ──', None, None, None, None, None, ''),
 (25, '목표 CAC (paid, ₩)', 30000, None, None, None, KRW, '근거: IR paid IG=30K. W5 실측으로 검증'),
 (27, '── 변동원가 (가구당/월) ──', None, None, None, None, None, ''),
 (28, '결제수수료율 (결제액 대비)', 0.033, None, None, None, PCT, 'PG 일반. 연매출 3억 미만 영세 우대 확정 시 ~1.5%로 하향'),
 (29, '인프라·스트리밍 (₩)', 1200, None, None, None, KRW, 'CDN egress+Supabase+GCS, 월 4시간 시청 가정'),
 (30, '개인화 L2 호명 등 (₩)', 100, None, None, None, KRW, 'TTS 이름 세그먼트'),
 (32, '── 콘텐츠 제작 (고정성 원가) ──', None, None, None, None, None, ''),
 (33, '에피소드 제작 원가 / 편 (₩)', 200000, None, None, None, KRW, '4~5분 에피소드, AI 파이프라인 기준(모델 스택 마스터플랜 참조)'),
 (34, '월 신규 에피소드 수', 8, None, None, None, INT, '근거: 백서 §8.2 소진 대응'),
 (36, '── B2G ──', None, None, None, None, None, ''),
 (37, '도서관 관당 연 라이선스 (공급가, ₩)', 18000000, None, None, None, KRW, 'Standard 티어. 수의계약 한도(소기업 5천만) 내 설계'),
 (39, '── 세율·부대율 ──', None, None, None, None, None, ''),
 (40, '부가가치세율', 0.10, None, None, None, PCT, ''),
 (41, '인건비 부대율 (4대보험+퇴직급여)', 0.19, None, None, None, PCT, '사업주 4대보험 ~10.5% + 퇴직급여 8.33%'),
 (43, '── 자금 ──', None, None, None, None, None, ''),
 (44, 'Seed 납입액 (₩, M1)', 300000000, None, None, None, KRW, 'IR: 3억 / post 30억 / 10%'),
]
for r, label, applied, cons, base, opt, fmt, note in rows:
    if applied is None and cons is None:
        put(A, f'A{r}', label, SEC); A[f'A{r}'].fill = FILL_SEC
        continue
    put(A, f'A{r}', label, BLACK)
    if isinstance(applied, str) and applied.startswith('='):
        put(A, f'B{r}', applied, BLACK, fmt)
        put(A, f'C{r}', cons, BLUE, fmt); put(A, f'D{r}', base, BLUE, fmt); put(A, f'E{r}', opt, BLUE, fmt)
    else:
        put(A, f'B{r}', applied, BLUE, fmt)
    if note: put(A, f'F{r}', note, Font(name='Arial', size=9, color='777777'))

put(A, 'A47', '── 인건비 계획 (역할 / 월급여 / 시작월) ──', SEC); A['A47'].fill = FILL_SEC
staff = [('CEO (최소화)', 3000000, 1), ('개발자 A', 4500000, 1), ('디자이너 (PT)', 2000000, 1),
         ('콘텐츠·검수', 3000000, 7), ('CS (PT)', 2000000, 10), ('개발자 B', 4500000, 13), ('마케터', 3500000, 13)]
put(A, 'A48', '역할', SEC); put(A, 'B48', '월급여(₩)', SEC); put(A, 'C48', '시작월', SEC)
for i, (role, sal, st) in enumerate(staff):
    r = 49 + i
    put(A, f'A{r}', role, BLACK); put(A, f'B{r}', sal, BLUE, KRW); put(A, f'C{r}', st, BLUE, INT)
put(A, 'A57', '월 기타 고정비 (임차·툴·법무회계, ₩)', BLACK); put(A, 'B57', 2500000, BLUE, KRW)

# ============ Sheet 2: 가입자 ============
S = wb.create_sheet('가입자')
S.column_dimensions['A'].width = 30; S.column_dimensions['B'].width = 8
for i in range(1, NM + 1): S.column_dimensions[mc(i)].width = 11
put(S, 'A1', '가입자·B2C 매출 모델 (M1 = 2026-07)', TITLE)

def month_header(ws, r3=3, r4=4):
    put(ws, f'A{r3}', '월 (M)', SEC); put(ws, f'A{r4}', '연월', SEC)
    y, m = 2026, 7
    for i in range(1, NM + 1):
        put(ws, f'{mc(i)}{r3}', i, SEC, INT, align='center')
        put(ws, f'{mc(i)}{r4}', f'{y}-{m:02d}', Font(name='Arial', size=9, color='777777'), align='center')
        m += 1
        if m == 13: m = 1; y += 1
month_header(S)

lib_ramp = [2,3,5,6,7,8,10,12,15,17,19,22,25,28,31,34,37,40,44,48,52,56,60,65,70,75,80,85,90,95,100,105,110,115,120,125]
mkt_ramp = [0,1000000,2000000,3000000,3000000,3000000,5000000,5000000,5000000,8000000,8000000,8000000] + [12000000]*6 + [20000000]*6 + [30000000]*12

def row(ws, r, label, unit, fmt, gen, font=BLACK, fill=None):
    put(ws, f'A{r}', label, SEC if fill else BLACK)
    if fill: ws[f'A{r}'].fill = fill
    put(ws, f'B{r}', unit, Font(name='Arial', size=9, color='777777'))
    for i in range(1, NM + 1):
        v = gen(i)
        if v is None: continue
        put(ws, f'{mc(i)}{r}', v, font, fmt)

row(S, 5, '운영 도서관 수 [입력]', '관', INT, lambda i: lib_ramp[i-1], BLUE)
row(S, 6, '키오스크 세션 수', '건', INT, lambda i: f"={mc(i)}5*'가정'!$B$17", BLACK)
row(S, 7, 'QR 스캔', '건', INT, lambda i: f"={mc(i)}6*'가정'!$B$18")
row(S, 8, '가입 완료', '가구', INT, lambda i: f"={mc(i)}7*'가정'!$B$19")
row(S, 9, '무료체험 시작', '가구', INT, lambda i: f"={mc(i)}8*'가정'!$B$20")
row(S, 10, '오가닉 신규 유료', '가구', INT, lambda i: f"={mc(i)}9*'가정'!$B$21")
row(S, 11, '마케팅 예산 [입력]', '₩', KRW, lambda i: mkt_ramp[i-1], BLUE)
row(S, 12, 'Paid 신규 유료', '가구', INT, lambda i: f"={mc(i)}11/'가정'!$B$25")
row(S, 13, '총 신규 유료', '가구', INT, lambda i: f"={mc(i)}10+{mc(i)}12")
row(S, 14, '이탈', '가구', INT, lambda i: 0 if i == 1 else f"={mc(i-1)}15*'가정'!$B$22")
row(S, 15, '기말 유료 가구', '가구', INT, lambda i: f"={mc(i)}13" if i == 1 else f"={mc(i-1)}15+{mc(i)}13-{mc(i)}14")
S[f'A15'].font = Font(name='Arial', size=10, bold=True)
row(S, 16, '연간플랜 신규', '가구', INT, lambda i: f"={mc(i)}13*'가정'!$B$13")
def trail12(i, r): 
    lo = max(1, i - 11); return f"=SUM({mc(lo)}{r}:{mc(i)}{r})"
row(S, 17, '연간플랜 활성', '가구', INT, lambda i: trail12(i, 16))
row(S, 18, '월간플랜 활성', '가구', INT, lambda i: f"={mc(i)}15-{mc(i)}17")
row(S, 19, '적용 월 ARPU (VAT포함)', '₩', KRW, lambda i: f"=IF({mc(i)}3<='가정'!$B$9,'가정'!$B$7,'가정'!$B$8*(1-'가정'!$B$10))")
row(S, 20, '형제 애드온 결제액 (VAT포함)', '₩', KRW, lambda i: f"={mc(i)}15*'가정'!$B$11*'가정'!$B$12")
row(S, 21, 'B2C 결제액 (VAT포함, 발생)', '₩', KRW, lambda i: f"={mc(i)}18*{mc(i)}19+{mc(i)}17*'가정'!$B$14/12+{mc(i)}20")
row(S, 22, 'B2C 순매출 (VAT제외)', '₩', KRW, lambda i: f"={mc(i)}21/(1+'가정'!$B$40)")
S['A22'].font = Font(name='Arial', size=10, bold=True)
row(S, 23, '(참고) 월 무료 체험자', '가구', INT, lambda i: f"={mc(i)}9")

# ============ Sheet 3: 손익 ============
P = wb.create_sheet('손익')
P.column_dimensions['A'].width = 30; P.column_dimensions['B'].width = 8
for i in range(1, NM + 1): P.column_dimensions[mc(i)].width = 11
put(P, 'A1', '월별 손익계산서 (발생주의, VAT 제외 순액)', TITLE)
month_header(P)

b2g_ramp = [0,0,1,1,2,2,3,3,4,4,5,5,6,7,7,8,9,10,11,12,13,14,15,16,18,19,21,22,24,25,26,27,28,29,30,32]
row(P, 6, 'B2C 순매출', '₩', KRW, lambda i: f"='가입자'!{mc(i)}22", GREEN)
row(P, 7, 'B2G 유료 관수 [입력]', '관', INT, lambda i: b2g_ramp[i-1], BLUE)
row(P, 8, 'B2G 순매출 (월할 인식)', '₩', KRW, lambda i: f"={mc(i)}7*'가정'!$B$37/12")
row(P, 9, '총매출', '₩', KRW, lambda i: f"={mc(i)}6+{mc(i)}8")
P['A9'].font = Font(name='Arial', size=10, bold=True)
row(P, 11, '결제수수료', '₩', KRW, lambda i: f"='가입자'!{mc(i)}21*'가정'!$B$28")
row(P, 12, '인프라·스트리밍', '₩', KRW, lambda i: f"='가입자'!{mc(i)}15*'가정'!$B$29")
row(P, 13, '개인화 COGS', '₩', KRW, lambda i: f"='가입자'!{mc(i)}15*'가정'!$B$30")
row(P, 14, '콘텐츠 제작비', '₩', KRW, lambda i: f"='가정'!$B$33*'가정'!$B$34")
row(P, 15, '매출원가 계', '₩', KRW, lambda i: f"=SUM({mc(i)}11:{mc(i)}14)")
row(P, 16, '매출총이익', '₩', KRW, lambda i: f"={mc(i)}9-{mc(i)}15")
P['A16'].font = Font(name='Arial', size=10, bold=True)
row(P, 17, '매출총이익률', '%', PCT, lambda i: f"=IF({mc(i)}9=0,0,{mc(i)}16/{mc(i)}9)")
row(P, 19, '인건비 (부대비 포함)', '₩', KRW, lambda i: f"=SUMPRODUCT(('가정'!$C$49:$C$55<={mc(i)}3)*'가정'!$B$49:$B$55)*(1+'가정'!$B$41)")
row(P, 20, '마케팅비', '₩', KRW, lambda i: f"='가입자'!{mc(i)}11", GREEN)
row(P, 21, '기타 고정비', '₩', KRW, lambda i: f"='가정'!$B$57")
row(P, 22, '판관비 계', '₩', KRW, lambda i: f"=SUM({mc(i)}19:{mc(i)}21)")
row(P, 24, '영업이익', '₩', KRW, lambda i: f"={mc(i)}16-{mc(i)}22")
P['A24'].font = Font(name='Arial', size=10, bold=True)
row(P, 25, '누적 영업이익', '₩', KRW, lambda i: f"={mc(i)}24" if i == 1 else f"={mc(i-1)}25+{mc(i)}24")
put(P, 'A27', '주: 법인세는 흑자 전환 전 0. 흑자 후 창업중소기업 세액감면(본점 소재지별 50~100%, 5년, 연 5억 한도) 적용 검토 — 문서 §7 참조', Font(name='Arial', size=9, italic=True, color='777777'))

# ============ Sheet 4: 현금 ============
Csh = wb.create_sheet('현금')
Csh.column_dimensions['A'].width = 30; Csh.column_dimensions['B'].width = 8
for i in range(1, NM + 1): Csh.column_dimensions[mc(i)].width = 11
put(Csh, 'A1', '현금흐름 · 런웨이 (VAT 포함 수금/지급 기준)', TITLE)
month_header(Csh)

gov_ramp = [0, 30000000] + [0] * (NM - 2)
row(Csh, 6, 'B2C 수금 (VAT포함)', '₩', KRW, lambda i: f"='가입자'!{mc(i)}18*'가입자'!{mc(i)}19+'가입자'!{mc(i)}20+'가입자'!{mc(i)}16*'가정'!$B$14", GREEN)
row(Csh, 7, 'B2G 수금 (VAT포함)', '₩', KRW, lambda i: f"='손익'!{mc(i)}8*(1+'가정'!$B$40)", GREEN)
row(Csh, 8, '정부지원금 [입력]', '₩', KRW, lambda i: gov_ramp[i-1], BLUE)
row(Csh, 9, 'Seed 납입', '₩', KRW, lambda i: "='가정'!$B$44" if i == 1 else 0)
row(Csh, 10, '유입 계', '₩', KRW, lambda i: f"=SUM({mc(i)}6:{mc(i)}9)")
row(Csh, 12, '매출원가 지급', '₩', KRW, lambda i: f"='손익'!{mc(i)}15", GREEN)
row(Csh, 13, '판관비 지급', '₩', KRW, lambda i: f"='손익'!{mc(i)}22", GREEN)
def vat_f(i):
    if i < 4 or (i - 1) % 3 != 0: return 0
    lo, hi = mc(i - 3), mc(i - 1)
    return (f"=MAX(0,(SUM('손익'!{lo}9:{hi}9)-SUM('손익'!{lo}14:{hi}14)"
            f"-SUM('손익'!{lo}20:{hi}20)-SUM('손익'!{lo}21:{hi}21))*'가정'!$B$40)")
row(Csh, 14, '부가세 순납부 (분기, 간이추정)', '₩', KRW, vat_f)
row(Csh, 15, '유출 계', '₩', KRW, lambda i: f"=SUM({mc(i)}12:{mc(i)}14)")
row(Csh, 17, '순현금흐름', '₩', KRW, lambda i: f"={mc(i)}10-{mc(i)}15")
row(Csh, 18, '기초 현금', '₩', KRW, lambda i: 0 if i == 1 else f"={mc(i-1)}19")
row(Csh, 19, '기말 현금', '₩', KRW, lambda i: f"={mc(i)}17" if i == 1 else f"={mc(i)}18+{mc(i)}17")
Csh['A19'].font = Font(name='Arial', size=10, bold=True)
def runway(i):
    lo = max(1, i - 2)
    avg = f"AVERAGE({mc(lo)}17:{mc(i)}17)"
    return f'=IF({mc(i)}19<=0,"소진",IF({avg}>=0,"흑자권",ROUND({mc(i)}19/ABS({avg}),1)))'
row(Csh, 20, '런웨이 (개월, 최근3M 소진 기준)', '개월', NUM, runway)
row(Csh, 21, '(메모) 선수수익 잔액', '₩', KRW, lambda i: (f"='가입자'!{mc(i)}16*'가정'!$B$14/(1+'가정'!$B$40)-'가입자'!{mc(i)}17*'가정'!$B$14/12/(1+'가정'!$B$40)" if i == 1 else f"={mc(i-1)}21+'가입자'!{mc(i)}16*'가정'!$B$14/(1+'가정'!$B$40)-'가입자'!{mc(i)}17*'가정'!$B$14/12/(1+'가정'!$B$40)"))
put(Csh, 'A23', '주: B2G는 발생=수금 가정(실제 계약 시 선수금 협상으로 개선 여지). 부가세는 인건비 매입세액 없음 반영한 간이 추정.', Font(name='Arial', size=9, italic=True, color='777777'))

# ============ Sheet 5: 유닛 ============
U = wb.create_sheet('유닛')
U.column_dimensions['A'].width = 36
for c in 'BCDEFG': U.column_dimensions[c].width = 15
put(U, 'A1', '유닛 이코노믹스 · 민감도 · BEP · 마일스톤', TITLE)

put(U, 'A3', 'A. 유닛 이코노믹스 정본 (정가 기준)', SEC); U['A3'].fill = FILL_SEC
ue = [
 ('정가 월 구독료 (VAT포함)', "='가정'!$B$8", KRW),
 ('순 ARPU (VAT제외)', "=B4/(1+'가정'!$B$40)", KRW),
 ('가구당 월 변동원가', "=B4*'가정'!$B$28+'가정'!$B$29+'가정'!$B$30", KRW),
 ('가구당 월 공헌이익', '=B5-B6', KRW),
 ('공헌마진율', '=B7/B5', PCT),
 ('월 이탈률 (시나리오)', "='가정'!$B$22", PCT),
 ('평균 구독기간 (개월)', '=1/B9', NUM),
 ('매출 LTV (순매출 기준)', '=B5*B10', KRW),
 ('공헌이익 LTV ★의사결정 기준', '=B7*B10', KRW),
 ('Paid CAC', "='가정'!$B$25", KRW),
 ('LTV/CAC (공헌이익 기준)', '=B12/B13', MULT),
 ('CAC 회수기간 (개월)', '=B13/B7', NUM),
 ('LTV/CAC 3.0x 허용 최대 CAC', '=B12/3', KRW),
]
for i, (label, f, fmt) in enumerate(ue):
    r = 4 + i
    put(U, f'A{r}', label, BLACK); put(U, f'B{r}', f, GREEN if '가정' in f else BLACK, fmt)
U['A12'].font = Font(name='Arial', size=10, bold=True); U['A15'].font = Font(name='Arial', size=10, bold=True)

put(U, 'A19', 'B. 공헌이익 LTV 민감도 (행=월이탈률, 열=월가격 VAT포함)', SEC); U['A19'].fill = FILL_SEC
prices = [19000, 25000, 29000]; churns = [0.15, 0.20, 0.25, 0.30, 0.35]
for j, p in enumerate(prices):
    put(U, f'{get_column_letter(3+j)}20', p, BLUE, KRW, align='center')
for i, ch in enumerate(churns):
    r = 21 + i
    put(U, f'B{r}', ch, BLUE, PCT)
    for j in range(3):
        col = get_column_letter(3 + j)
        f = f"=({col}$20/(1+'가정'!$B$40)-({col}$20*'가정'!$B$28+'가정'!$B$29+'가정'!$B$30))/$B{r}"
        put(U, f'{col}{r}', f, BLACK, KRW)

put(U, 'A28', 'C. 손익분기점 (BEP)', SEC); U['A28'].fill = FILL_SEC
bep = [
 ('월 고정비 — 팀 급여 포함 (M12 기준)', f"='손익'!{mc(12)}19+'손익'!{mc(12)}21+'손익'!{mc(12)}14", KRW),
 ('BEP 유료 가구 (팀 포함)', '=B29/B7', INT),
 ('월 고정비 — 운영 최소 (인건비 제외)', f"='손익'!{mc(12)}21+'손익'!{mc(12)}14", KRW),
 ('BEP 유료 가구 (운영 최소)', '=B31/B7', INT),
]
for i, (label, f, fmt) in enumerate(bep):
    r = 29 + i
    put(U, f'A{r}', label, BLACK); put(U, f'B{r}', f, GREEN, fmt)
put(U, 'A33', '검증: 백서 v1.1 §8.3 = 운영최소 200명 / 팀포함 800명 (인건비 1,200만 가정) — 본 모델과 정합 확인용', Font(name='Arial', size=9, italic=True, color='777777'))

put(U, 'A35', 'D. 마일스톤 체크 (IR 로드맵 대비)', SEC); U['A35'].fill = FILL_SEC
put(U, 'A36', '마일스톤', SEC); put(U, 'B36', '목표', SEC); put(U, 'C36', '모델값', SEC); put(U, 'D36', '판정', SEC); put(U, 'E36', '갭', SEC)
ms = [('M+2 유료 가구 (스프린트)', 20, 2), ('M+9 유료 가구 (Pre-A 트리거)', 1000, 9), ('M+18 유료 가구 (Series A)', 10000, 18)]
for i, (label, tgt, m) in enumerate(ms):
    r = 37 + i
    put(U, f'A{r}', label, BLACK); put(U, f'B{r}', tgt, BLUE, INT)
    put(U, f'C{r}', f"='가입자'!{mc(m)}15", GREEN, INT)
    put(U, f'D{r}', f'=IF(C{r}>=B{r},"달성","미달")', BLACK)
    put(U, f'E{r}', f'=C{r}-B{r}', BLACK, INT)
put(U, 'A41', '월 MRR (M+9, 순매출)', BLACK); put(U, 'B41', f"='가입자'!{mc(9)}22", GREEN, KRW)
put(U, 'A42', 'ARR 런레이트 (M+12, B2C+B2G)', BLACK); put(U, 'B42', f"='손익'!{mc(12)}9*12", GREEN, KRW)

# [macOS patch] Write generated artifacts under the repo-local simulation output directory.
wb.save(OUT_DIR / 'Kindy_재무모델_v1.0.xlsx')
print('saved')
